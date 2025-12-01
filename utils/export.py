from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
from config import Config

class ReportExporter:
    """Export predictions to PDF and Excel"""
    
    def __init__(self):
        os.makedirs(Config.EXPORT_DIR, exist_ok=True)
    
    def export_to_pdf(self, predictions, user_info, filename=None):
        """Export predictions to PDF"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"predictions_{timestamp}.pdf"
        
        filepath = os.path.join(Config.EXPORT_DIR, filename)
        
        # Create PDF
        doc = SimpleDocTemplate(filepath, pagesize=A4,
                               rightMargin=0.5*inch, leftMargin=0.5*inch,
                               topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Container for elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#6366f1'),
            spaceAfter=12
        )
        
        # Title
        title = Paragraph("College Prediction Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # User Information
        user_data = [
            ['Student Name:', user_info.get('name', 'N/A')],
            ['Email:', user_info.get('email', 'N/A')],
            ['Report Date:', datetime.now().strftime("%B %d, %Y")],
            ['Total Predictions:', str(len(predictions))]
        ]
        
        user_table = Table(user_data, colWidths=[2*inch, 4*inch])
        user_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        
        elements.append(user_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Predictions heading
        pred_heading = Paragraph("Predicted Colleges", heading_style)
        elements.append(pred_heading)
        elements.append(Spacer(1, 0.1*inch))
        
        # Predictions table
        table_data = [['S.No', 'College', 'Program', 'Degree', 'Opening Rank', 'Closing Rank', 'Match %']]
        
        for idx, pred in enumerate(predictions, 1):
            table_data.append([
                str(idx),
                pred.get('college', 'N/A'),
                pred.get('program', 'N/A')[:30] + '...' if len(pred.get('program', '')) > 30 else pred.get('program', 'N/A'),
                pred.get('degree', 'N/A'),
                str(pred.get('opening_rank', 'N/A')),
                str(pred.get('closing_rank', 'N/A')),
                f"{pred.get('match_score', 0):.1f}%"
            ])
        
        pred_table = Table(table_data, colWidths=[0.4*inch, 1.2*inch, 1.8*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.6*inch])
        pred_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(pred_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Footer note
        note_style = ParagraphStyle(
            'Note',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        note = Paragraph(
            "Note: This prediction is based on historical data and actual admissions may vary. "
            "Please verify with official sources before making decisions.",
            note_style
        )
        elements.append(note)
        
        # Build PDF
        doc.build(elements)
        
        return filepath
    
    def export_to_excel(self, predictions, user_info, filename=None):
        """Export predictions to Excel"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"predictions_{timestamp}.xlsx"
        
        filepath = os.path.join(Config.EXPORT_DIR, filename)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "College Predictions"
        
        # Styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        cell_alignment = Alignment(horizontal='left', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = 'College Prediction Report'
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='1e40af')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # User Info
        ws['A3'] = 'Student Name:'
        ws['B3'] = user_info.get('name', 'N/A')
        ws['A4'] = 'Email:'
        ws['B4'] = user_info.get('email', 'N/A')
        ws['A5'] = 'Report Date:'
        ws['B5'] = datetime.now().strftime("%B %d, %Y")
        
        for row in range(3, 6):
            ws[f'A{row}'].font = Font(bold=True)
        
        # Headers
        headers = ['S.No', 'College', 'Program', 'Degree', 'Opening Rank', 'Closing Rank', 'Match Score (%)']
        ws.append([])  # Empty row
        ws.append(headers)
        
        header_row = ws.max_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data
        for idx, pred in enumerate(predictions, 1):
            row = [
                idx,
                pred.get('college', 'N/A'),
                pred.get('program', 'N/A'),
                pred.get('degree', 'N/A'),
                pred.get('opening_rank', 'N/A'),
                pred.get('closing_rank', 'N/A'),
                f"{pred.get('match_score', 0):.1f}"
            ]
            ws.append(row)
            
            # Apply styling
            current_row = ws.max_row
            for col_num in range(1, 8):
                cell = ws.cell(row=current_row, column=col_num)
                cell.alignment = cell_alignment
                cell.border = border
        
        # Adjust column widths
        column_widths = [8, 25, 40, 12, 15, 15, 15]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width
        
        # Save
        wb.save(filepath)
        
        return filepath
    
    def export_admin_report(self, statistics, users, predictions):
        """Export admin dashboard report"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"admin_report_{timestamp}.xlsx"
        filepath = os.path.join(Config.EXPORT_DIR, filename)
        
        wb = Workbook()
        
        # Statistics sheet
        ws1 = wb.active
        ws1.title = "Statistics"
        
        ws1['A1'] = 'System Statistics Report'
        ws1['A1'].font = Font(size=16, bold=True)
        ws1.merge_cells('A1:B1')
        
        ws1['A3'] = 'Total Users'
        ws1['B3'] = statistics.get('total_users', 0)
        ws1['A4'] = 'Total Predictions'
        ws1['B4'] = statistics.get('total_predictions', 0)
        ws1['A5'] = 'Total Colleges'
        ws1['B5'] = statistics.get('total_colleges', 0)
        ws1['A6'] = 'Recent Users (7 days)'
        ws1['B6'] = statistics.get('recent_users', 0)
        
        # Users sheet
        ws2 = wb.create_sheet("Users")
        ws2.append(['ID', 'Name', 'Email', 'Phone', 'Created At', 'Last Login'])
        
        for user in users:
            ws2.append([
                user.get('id'),
                user.get('name'),
                user.get('email'),
                user.get('phone', 'N/A'),
                user.get('created_at'),
                user.get('last_login', 'Never')
            ])
        
        # Predictions sheet
        ws3 = wb.create_sheet("Recent Predictions")
        ws3.append(['ID', 'User Email', 'Rank', 'Category', 'Year', 'Date'])
        
        for pred in predictions[:100]:  # Last 100 predictions
            ws3.append([
                pred.get('id'),
                pred.get('email', 'Guest'),
                pred.get('rank'),
                pred.get('category'),
                pred.get('year'),
                pred.get('prediction_date')
            ])
        
        wb.save(filepath)
        return filepath